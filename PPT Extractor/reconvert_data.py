import subprocess
import pandas as pd
import numpy as np
from ttkbootstrap.dialogs import Messagebox
from openpyxl import load_workbook

class ReConvert:
    def __init__(self, root, dtype, progress, progress_var, frame, lbl_status, projectpath):
        self.root = root
        self.dtype = dtype
        self.progress = progress
        self.progress_var = progress_var
        self.frame = frame
        self.lbl_status = lbl_status
        self.projectpath = f'projects/{projectpath.get()}'

    def reconvert(self):
        cnv = f'{self.projectpath}/output/converted_data.xlsx'
        out = f'{self.projectpath}/output/output_data.xlsx'
        rec_all = self.reconvert_all(cnv)
        self.save_to_excel(out, rec_all)

        result = Messagebox.yesno(title="Options",message="Your data has been successfully extracted to "+out+"\n\nDo you want to open the output?",parent=self.frame)
        if result == 'Yes':
            subprocess.run(['start', 'excel', out], shell=True)

        self.progress.pack_forget()
        self.lbl_status.config(text='')

    def reconvert_texts(self, file_converted):
        self.lbl_status.config(text="Reconverting Texts")
        wb = load_workbook(file_converted)
        for item in wb:
            grouped_text = {}
            n = False
            for index, row in enumerate(item.values, start=1):
                self.progress_var.set(index / len(list(item.values)) * 100)
                self.root.update_idletasks()
                if index != 1:
                    if row[11] == "Text":
                        key = "{}#{}#{}".format(f'Slide {row[0]}', row[1], row[2])
                        if key not in grouped_text:
                            grouped_text[key] = []

                        if row[7]:
                            grouped_text[key].append(row[7])
                        else:
                            grouped_text[key].append(row[6])
                    
            dataframes = [{"Identity" : i, "Data": d} for i, d in grouped_text.items()]
            
            return dataframes

    def reconvert_tables(self, file_converted):
        self.lbl_status.config(text="Reconverting Tables")
        wb = load_workbook(file_converted)

        for item in wb:
            grouped_tables = {}
            for index, value in enumerate(item.values):
                if value[11] == 'Table':
                    self.progress_var.set(index / len(list(item.values)) * 100)
                    self.root.update_idletasks()
                    val_selected = 7 if value[7] else 6
                    filtered = [3,4,5,val_selected,12]
                    contents = tuple(value[i] for i in filtered)
                    key = "{}#{}#{}".format(f'Slide {value[0]}',value[1],value[2])
                    if key not in grouped_tables:
                        grouped_tables[key] = []
                    grouped_tables[key].append(contents)
                
            dataframes = [{"Identity" : i, "Data": d} for i, d in grouped_tables.items()]

            return dataframes

    def reconvert_charts(self, file_converted):
        self.lbl_status.config(text="Reconverting Charts")
        wb = load_workbook(file_converted)
        for item in wb:
            grouped_charts = {}
            for index, value in enumerate(item.values):
                if value[11] == "Chart":
                    self.progress_var.set(index / len(list(item.values)) * 100)
                    self.root.update_idletasks()
                    if index:
                        val_selected = 7 if value[7] else 6
                        filtered = [4,5,val_selected]
                        contents = tuple(value[i] for i in filtered)
                        key = "{}#{}#{}#{}".format(f'Slide {value[0]}',value[1],value[2],value[12])
                        
                        if key not in grouped_charts:
                            grouped_charts[key] = []
                        grouped_charts[key].append(contents)
                    
            dataframes = [{"Identity" : i, "Data": d} for i, d in grouped_charts.items()]
            return dataframes

    def reconvert_all(self, file_converted):
        all_data = []
        if 'Texts' in self.dtype:
            all_data.append(self.reconvert_texts(file_converted))
        if 'Tables' in self.dtype:
            all_data.append(self.reconvert_tables(file_converted))
        if 'Charts' in self.dtype:
            all_data.append(self.reconvert_charts(file_converted))

        return all_data

    def save_to_excel(self, file_output, dataframes):
        with pd.ExcelWriter(file_output, engine='xlsxwriter') as writer:
            for dtype, list_dfs in zip(self.dtype, dataframes):
                if dtype == 'Texts':
                    for item in list_dfs:
                        item['Data'] = pd.DataFrame({'Text': item['Data']})
                
                    row_usage = 1
                    sh_name = 'Texts'
                    for index, df in enumerate(list_dfs, start=1):
                        df['Data'].to_excel(writer, sheet_name=sh_name, index=False, startrow=row_usage-1, startcol=1, header=True)
                        worksheet = writer.sheets[sh_name]
                        worksheet.write_string(row_usage-1, 0, df['Identity'])
                        row_usage += len(df['Data'])+2
                
                if dtype == 'Tables':
                    row_usage = 1
                    sh_name = 'Tables'
                    for i, dfs in enumerate(list_dfs,start=1):
                        col = dfs['Data'][0][4]
                        df = pd.DataFrame(dfs['Data'],columns=("ID","Header",col,"Values","Info"))
                        df = df.iloc[:,:4]
                        sort_col = np.array(df.iloc[:,1].unique())
                        sort_col = np.insert(sort_col, 0, col)
                        sort_col = [x for x in sort_col if x != "EmptyValues"]
                        pivoted_df = df.pivot(index=["ID",col], columns='Header', values='Values',).reset_index()
                        pivoted_df = pivoted_df.iloc[:,1:]
                        
                        new_df = pivoted_df.loc[:, pivoted_df.columns != 'EmptyValues']
                        new_df = new_df[sort_col]

                        unnamed_columns = [col for col in new_df.columns if col.startswith('Unnamed')]

                        if unnamed_columns:
                            new_columns = {col: '' for col in unnamed_columns}
                            new_df.rename(columns=new_columns, inplace=True)
                        
                        new_df.to_excel(writer, sheet_name=sh_name, index=False, startrow=row_usage-1, startcol=1, header=True)
                        
                        worksheet = writer.sheets[sh_name]
                        worksheet.write_string(row_usage-1, 0, dfs['Identity']) 
                        row_usage += len(pivoted_df) + 2

                if dtype == 'Charts':
                    row_usage = 1
                    sh_name = 'Charts'
                    for i, dfs in enumerate(list_dfs,start=1):
                        df = pd.DataFrame(dfs['Data'],columns=("Legend","Categories","Values"))
                        df.to_excel(writer, sheet_name=sh_name, index=False, startrow=row_usage-1, startcol=1, header=True)

                        worksheet = writer.sheets[sh_name]
                        worksheet.write_string(row_usage-1, 0, dfs['Identity']) 
                        row_usage += len(dfs['Data']) + 2
