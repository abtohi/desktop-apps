import pandas as pd
import subprocess
from openpyxl import load_workbook
from ttkbootstrap.dialogs import Messagebox
#ngetes
class ConvertData:
    def __init__(self, **kwargs):
        self.lbl_status = kwargs['lbl_status']
        self.dtype = kwargs['dtype']
        self.root = kwargs['root']
        self.frame4 = kwargs['frame4']
        self.projectpath = f"projects\{kwargs['projectpath'].get()}"

    def convert(self, progress_var, progress):
        try:
            progress_var.set(1)
            df = self.convert_all(progress_var)
            out = f'{self.projectpath}\output\converted_data.xlsx'
            
            self.save_to_excel(df,out)
            
            self.lbl_status.config(text="You have successfully converted the format of your data table")
            result = Messagebox.yesno("Your data table format has been successfully converted, do you want to open the file?", parent=self.frame4)
            if result == 'Yes':
                subprocess.run(['start', 'excel', out], shell=True)
                progress.pack_forget()
                self.lbl_status.config(text="")
            else:
                progress.pack_forget()
                self.lbl_status.config(text="")
        except Exception as e:
            error_msg = str(e)
            Messagebox.show_error(title="Error",message=error_msg,parent=self.frame4)
            progress.pack_forget()

    def convert_texts(self, progress_var):
        self.lbl_status.config(text="Converting Texts")
        filename = f"{self.projectpath}\output\output_data.xlsx"
        sh_name = "Texts"
        data = self.read_texts_out(filename, sh_name, progress_var, 0, 100)
                    
        out = []
        for df in data:
            key = df['Identity']
            value = [item for item in df['Data'] if item is not None and item != 'Text']
            for i, text in enumerate(value, start=1):
                slide = int(key.split('#')[0].replace('Slide ',''))
                obj = key.split('#')[1]
                size = key.split('#')[2]
                out.append([slide,obj,size,i,"Text","Text",text.lstrip('0123456789').strip(),"","","","","Text",""])
        
        columns = ["Slide","Object","Size","Index","Header","Categories","PPT Value","Revised Value","Key Question","Key Response","Key Header","Type","Info"]
        final_df = pd.DataFrame(out, columns=columns)
        return final_df

    def convert_tables(self, progress_var):
        self.lbl_status.config(text="Converting Tables")
        filename = f"{self.projectpath}\output\output_data.xlsx"
        sh_name = "Tables"
        wb = load_workbook(filename)

        data = []
        for item in wb:
            value_list = list(item.values)
            if sh_name == item.title:
                for index, value in enumerate(item.values, start=1):
                    progress_var.set((index / len(value_list)) * 100)
                    self.root.update_idletasks()
                    nrow = 0
                    ncol = 0
                    if value[0] != None:
                        shape = str(value[0]).split('#')[2]
                        nrow = int(str(shape).split('x')[0])
                        ncol = int(str(shape).split('x')[1])

                    if nrow != 0 and ncol != 0:
                        df = pd.read_excel(filename, nrows=nrow, sheet_name=sh_name, skiprows=index-1, usecols=range(1,ncol+1), index_col=None)
                        df_reset = df.reset_index()
                        df_reset.rename(columns={'index': 'ID'}, inplace=True)
                        val = {'Identity':value[0], 'Data':df_reset}
                        data.append(val)
        
        out = []
        for df in data:
            key = df['Identity']
            value = df['Data'].fillna('')
            info = value.columns[1]
            slide = int(key.split('#')[0].replace('Slide ',''))
            obj = key.split('#')[1]
            size = key.split('#')[2]
            id_vars = [value.columns[0],value.columns[1]]
            try:
                value.iloc[:,2]
            except:
                value['EmptyValues'] = 'EmptyValues'
            melted_df = pd.melt(value, id_vars=id_vars, var_name='Header', value_name='Value')

            melted_df.rename(columns={value.columns[1]: 'Categories'}, inplace=True)
            for index, (id, header, category, val) in enumerate(zip(melted_df['ID'],melted_df['Header'], melted_df['Categories'],melted_df['Value']), start=1):
                out.append([slide,obj,size,id, header,category,val,"","","","","Table",info])

        columns = ["Slide","Object","Size","Index","Header","Categories","PPT Value","Revised Value","Key Question","Key Response","Key Header","Type","Info"]
        final_df = pd.DataFrame(out, columns=columns)
        return final_df

    def convert_charts(self, progress_var):
        self.lbl_status.config(text="Converting Charts")
        filename = f"{self.projectpath}\output\output_data.xlsx"
        sh_name = "Charts"

        wb = load_workbook(filename)

        data = []
        for item in wb:
            value_list = list(item.values)
            if sh_name == item.title:
                for index, value in enumerate(item.values, start=1):
                    progress_var.set(index / len(value_list) * 100)
                    self.root.update_idletasks()
                    nrow = 0
                    ncol = 0
                    if value[0] != None:
                        shape = str(value[0]).split('#')[2]
                        nrow = int(str(shape).split('x')[0])
                        ncol = int(str(shape).split('x')[1])

                    if nrow != 0 and ncol != 0:
                        df = pd.read_excel(filename, nrows=nrow, sheet_name=sh_name, skiprows=index-1, usecols=range(1,ncol+1), index_col=None)
                        val = {'Identity':value[0], 'Data':df}
                        data.append(val)
                    
        out = []
        for df in data:
            key = df['Identity']
            value = df['Data'].fillna('')
            slide = int(key.split('#')[0].replace('Slide ',''))
            obj = key.split('#')[1]
            size = key.split('#')[2]
            chart_type = key.split('#')[3]
            for index, (header, category, val) in enumerate(zip(value['Legend'],value['Categories'],value['Values']), start=1):
                out.append([slide,obj,size,index,header,category,val,"","","","","Chart",chart_type])

        columns = ["Slide","Object","Size","Index","Header","Categories","PPT Value","Revised Value","Key Question","Key Response","Key Header","Type","Info"]
        final_df = pd.DataFrame(out, columns=columns)
        return final_df

    def convert_all(self, progress_var):
        df_gabung = pd.DataFrame()
        if 'Texts' in self.dtype:
            texts = self.convert_texts(progress_var)
            df_gabung = pd.concat([df_gabung,texts], axis=0, ignore_index=True)
        if 'Tables' in self.dtype:
            tables = self.convert_tables(progress_var)
            df_gabung = pd.concat([df_gabung, tables], axis=0, ignore_index=True)
        if 'Charts' in self.dtype:
            charts = self.convert_charts(progress_var)
            df_gabung = pd.concat([df_gabung, charts], axis=0, ignore_index=True)

        return df_gabung
    
    def save_to_excel(self, df, outputname):
        with pd.ExcelWriter(outputname, engine='openpyxl') as writer:
            df.to_excel(writer,sheet_name="Data",index=False)
            workbook = writer.book
            worksheet = writer.sheets['Data']
            worksheet.freeze_panes = 'A2'

    def read_texts_out(self, filename, sheetname, progress_var, cons, mult):
        wb = load_workbook(filename)
        for item in wb:
            value_list = list(item.values)
            if item.title == sheetname:
                grouped_data = {}
                for index, (key, val) in enumerate(item.values, start=1):
                    progress_var.set(cons + ((index / len(value_list)) * mult))
                    self.root.update_idletasks()
                    if key:
                        keys = key                 
                    if keys not in grouped_data:
                        grouped_data[keys] = []
                    grouped_data[keys].append(val)
            
        output = [{"Identity" : i, "Data": d} for i, d in grouped_data.items()]
        return output